#!/usr/bin/env python3
"""Importa a catalogação ABNT NBR 6023 do cliente para o catálogo do Guapu.

Entrada: a planilha `referencias_enfermagem_perioperatoria.xlsx`, com as colunas
`Pasta`, `Arquivo` e `Referência (ABNT NBR 6023)`.

O vínculo é feito pelo nome do arquivo, que segue a mesma convenção técnica do
inventário do Drive (`tema__subtema__tipo__autor__ano__vN`). O script nunca
inventa um vínculo: uma linha que não casar com nenhum documento conhecido é
listada no relatório para conferência manual, e nada é gravado para ela.

Saídas:
  - `reference_catalog.json`               (bootstrap consumido pela ingestão)
  - `lib/chat/document-catalog.ts`         (bootstrap consumido pelo app)
  - `db/seeds/abnt_reference_catalog.sql`  (atualização da tabela no Supabase)
  - `scratch/abnt_import_report.md`        (casados, não casados, divergências)

Uso:
    python scripts/import_abnt_references.py caminho/para/planilha.xlsx
    python scripts/import_abnt_references.py planilha.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / 'scratch' / 'catalogacao_raw_20260830.json'
CATALOG_JSON = ROOT / 'reference_catalog.json'
CATALOG_TS = ROOT / 'lib' / 'chat' / 'document-catalog.ts'
SEED_SQL = ROOT / 'db' / 'seeds' / 'abnt_reference_catalog.sql'
REPORT = ROOT / 'scratch' / 'abnt_import_report.md'

COLUMN_ALIASES = {
    'pasta': 'folder',
    'arquivo': 'file',
    'referencia': 'reference',
    'referencia abnt nbr 6023': 'reference',
    'referencia abnt': 'reference',
}


def normalize(value: str) -> str:
    """Chave de comparação: sem acentos, sem extensão, sem separadores."""
    text = unicodedata.normalize('NFD', str(value or ''))
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = text.lower().strip()
    text = re.sub(r'\.(pdf|docx?|pptx?|txt)$', '', text)
    return re.sub(r'[^a-z0-9]+', '', text)


def load_manifest() -> dict[str, str]:
    if not MANIFEST.exists():
        sys.exit(f'Inventário não encontrado: {MANIFEST}')
    data = json.loads(MANIFEST.read_text(encoding='utf-8'))
    mapping: dict[str, str] = {}
    for item in data.get('items', []):
        name = item.get('inventory_name')
        file_id = item.get('drive_file_id')
        if name and file_id:
            mapping[normalize(name)] = file_id
    return mapping


def read_sheet(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        sys.exit('Planilha vazia.')

    columns: dict[int, str] = {}
    for index, cell in enumerate(header):
        key = normalize(cell or '').replace('nbr6023', '').strip()
        raw = re.sub(r'[^a-z0-9 ]+', ' ', unicodedata.normalize('NFKD', str(cell or ''))
                     .encode('ascii', 'ignore').decode().lower()).strip()
        raw = re.sub(r'\s+', ' ', raw)
        if raw in COLUMN_ALIASES:
            columns[index] = COLUMN_ALIASES[raw]
        elif 'arquivo' in raw:
            columns[index] = 'file'
        elif 'refer' in raw:
            columns[index] = 'reference'
        elif 'pasta' in raw:
            columns[index] = 'folder'
        elif key:
            columns[index] = key

    if 'file' not in columns.values() or 'reference' not in columns.values():
        sys.exit(f'Colunas Arquivo/Referência não encontradas. Cabeçalho lido: {header}')

    records = []
    for row in rows:
        record = {columns[i]: (row[i] if i < len(row) else None) for i in columns}
        if not str(record.get('file') or '').strip():
            continue
        records.append({
            'folder': str(record.get('folder') or '').strip(),
            'file': str(record.get('file') or '').strip(),
            'reference': ' '.join(str(record.get('reference') or '').split()),
        })
    return records


def ts_escape(value: str) -> str:
    return value.replace('\\', '\\\\').replace("'", "\\'")


def update_typescript(assignments: dict[str, str]) -> int:
    """Insere `reference_abnt` em cada entrada do bootstrap TypeScript.

    O fim de cada objeto é localizado por contagem de chaves, o que funciona
    tanto nas entradas de várias linhas quanto nas de uma linha só. Uma citação
    anterior do mesmo documento é substituída, nunca duplicada.
    """
    source = CATALOG_TS.read_text(encoding='utf-8')
    previous = re.compile(r"\s*reference_abnt:\s*'(?:[^'\\]|\\.)*',")
    applied = 0
    for file_id, reference in assignments.items():
        anchor = f"'{file_id}': {{"
        index = source.find(anchor)
        if index < 0:
            continue
        start = index + len(anchor)
        depth, cursor = 1, start
        while cursor < len(source) and depth:
            if source[cursor] == '{':
                depth += 1
            elif source[cursor] == '}':
                depth -= 1
            cursor += 1
        body = previous.sub('', source[start:cursor - 1], count=1)
        source = source[:start] + f"\n    reference_abnt: '{ts_escape(reference)}'," + body + source[cursor - 1:]
        applied += 1
    CATALOG_TS.write_text(source, encoding='utf-8')
    return applied


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('planilha', type=Path)
    parser.add_argument('--dry-run', action='store_true', help='Só relata; não grava nada.')
    args = parser.parse_args()

    if not args.planilha.exists():
        sys.exit(f'Planilha não encontrada: {args.planilha}')

    manifest = load_manifest()
    catalog = json.loads(CATALOG_JSON.read_text(encoding='utf-8'))
    rows = read_sheet(args.planilha)

    matched: dict[str, dict[str, str]] = {}
    unmatched: list[dict[str, str]] = []
    empty: list[dict[str, str]] = []

    for row in rows:
        if not row['reference']:
            empty.append(row)
            continue
        file_id = manifest.get(normalize(row['file']))
        if not file_id:
            unmatched.append(row)
            continue
        matched[file_id] = row

    not_in_sheet = [fid for fid in catalog if fid not in matched]

    lines = [
        '# Importação da catalogação ABNT — relatório',
        '',
        f'- Planilha: `{args.planilha.name}`',
        f'- Linhas lidas: {len(rows)}',
        f'- Documentos vinculados: {len(matched)} de {len(catalog)} no catálogo',
        f'- Linhas sem correspondência no inventário: {len(unmatched)}',
        f'- Linhas com referência em branco: {len(empty)}',
        f'- Documentos do catálogo sem linha na planilha: {len(not_in_sheet)}',
        '',
    ]
    if unmatched:
        lines += ['## Sem correspondência no inventário do Drive', '',
                  '| Pasta | Arquivo |', '| --- | --- |']
        lines += [f"| {r['folder']} | `{r['file']}` |" for r in unmatched] + ['']
    if not_in_sheet:
        lines += ['## No catálogo, mas ausentes da planilha', '',
                  '| Drive file ID | Título atual |', '| --- | --- |']
        lines += [f"| `{fid}` | {catalog[fid].get('reference_title', '')} |" for fid in not_in_sheet] + ['']

    if args.dry_run:
        print('\n'.join(lines))
        return

    for file_id, row in matched.items():
        entry = catalog.setdefault(file_id, {
            'reference_source': 'catalog', 'reference_verified': True, 'reference_key': file_id,
        })
        entry['reference_abnt'] = row['reference']

    CATALOG_JSON.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    applied = update_typescript({fid: row['reference'] for fid, row in matched.items()})

    SEED_SQL.parent.mkdir(parents=True, exist_ok=True)
    statements = ['-- Gerado por scripts/import_abnt_references.py — não editar à mão.',
                  '-- Aplicar depois da migração 043.', 'BEGIN;']
    for file_id, row in matched.items():
        value = row['reference'].replace("'", "''")
        statements.append(
            "UPDATE rag_document_catalog SET reference_abnt = '"
            f"{value}' WHERE drive_file_id = '{file_id}';")
    statements += ['SELECT sync_abnt_reference_to_documents();', 'COMMIT;', '']
    SEED_SQL.write_text('\n'.join(statements), encoding='utf-8')

    lines += ['## Resultado', '',
              f'- `reference_catalog.json`: {len(matched)} citações gravadas.',
              f'- `lib/chat/document-catalog.ts`: {applied} entradas atualizadas.',
              f'- `{SEED_SQL.relative_to(ROOT)}`: {len(matched)} UPDATEs + sincronização dos chunks.', '']
    REPORT.write_text('\n'.join(lines), encoding='utf-8')
    print('\n'.join(lines))


if __name__ == '__main__':
    main()
